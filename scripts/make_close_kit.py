#!/usr/bin/env python3
"""Build the SME Monthly Close Kit workbook.

A bookkeeping tool for running a small business's monthly close:
paste transactions in, categorize them from a dropdown, and the P&L,
cash summary and a client-ready report fill themselves in.

    python3 scripts/make_close_kit.py   -> dist/SME-Monthly-Close-Kit.xlsx
"""

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, "dist", "SME-Monthly-Close-Kit.xlsx")

TXN_ROWS = 500  # transaction capacity per month

ARIAL = "Arial"
INPUT_BLUE = Font(name=ARIAL, color="0000FF")
YELLOW = PatternFill("solid", fgColor="FFFF00")
HEAD_FILL = PatternFill("solid", fgColor="1F4E4B")
HEAD_FONT = Font(name=ARIAL, bold=True, color="FFFFFF")
BOLD = Font(name=ARIAL, bold=True)
PLAIN = Font(name=ARIAL)
GREEN = Font(name=ARIAL, color="008000")  # cross-sheet links
THIN = Border(*[Side(style="thin", color="CCCCCC")] * 4)

RM = '"RM"#,##0.00;("RM"#,##0.00);-'

INCOME_CATS = ["Sales", "Other income"]
EXPENSE_CATS = [
    "Purchases / stock", "Rent", "Salaries & EPF/SOCSO", "Utilities",
    "Marketing", "Transport & delivery", "Supplies & packaging",
    "Professional fees", "Loan repayment", "Other expenses",
]
CATS = INCOME_CATS + EXPENSE_CATS


def style(ws, cell, value, font=None, fill=None, fmt=None, border=False):
    c = ws[cell]
    c.value = value
    c.font = font or PLAIN
    if fill:
        c.fill = fill
    if fmt:
        c.number_format = fmt
    if border:
        c.border = THIN
    return c


def header_row(ws, row, headers, widths):
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        col = get_column_letter(i)
        style(ws, f"{col}{row}", h, font=HEAD_FONT, fill=HEAD_FILL, border=True)
        ws.column_dimensions[col].width = w


def main():
    os.makedirs(os.path.join(ROOT, "dist"), exist_ok=True)
    wb = Workbook()

    # ---------------------------------------------------------- READ ME
    ws = wb.active
    ws.title = "READ ME"
    ws.column_dimensions["A"].width = 100
    lines = [
        ("SME Monthly Close Kit", BOLD),
        ("One workbook per client per month. Save a copy named e.g. ClientName-2026-08.xlsx", PLAIN),
        ("", PLAIN),
        ("HOW TO USE", BOLD),
        ("1. Settings sheet: fill the yellow cells (client, month, opening cash).", PLAIN),
        ("2. Transactions sheet: enter every transaction for the month — date,", PLAIN),
        ("   description, money in OR money out, and pick a category from the dropdown.", PLAIN),
        ("3. That's it. P&L, Cash Summary and Client Report calculate themselves.", PLAIN),
        ("4. Check the Client Report page, then send it (or export it as PDF).", PLAIN),
        ("", PLAIN),
        ("COLOR LEGEND", BOLD),
        ("Yellow cells = you fill these in.  Blue text = typed inputs.", PLAIN),
        ("Black = formulas (do not overwrite).  Green = pulled from another sheet.", PLAIN),
        ("", PLAIN),
        ("NOTES", BOLD),
        ("Categories can be renamed on the Settings sheet — the whole book follows.", PLAIN),
        ("Row capacity: %d transactions. The one example row shows the format —" % TXN_ROWS, PLAIN),
        ("delete it before entering a real month.", PLAIN),
    ]
    for i, (text, font) in enumerate(lines, start=1):
        style(ws, f"A{i}", text, font=font)

    # ---------------------------------------------------------- Settings
    ws = wb.create_sheet("Settings")
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 30
    style(ws, "A1", "CLIENT SETTINGS", font=HEAD_FONT, fill=HEAD_FILL)
    style(ws, "A2", "Client / business name")
    style(ws, "B2", "Example Trading Sdn Bhd", font=INPUT_BLUE, fill=YELLOW)
    style(ws, "A3", "Month")
    style(ws, "B3", "August 2026", font=INPUT_BLUE, fill=YELLOW)
    style(ws, "A4", "Opening cash balance")
    style(ws, "B4", 5000, font=INPUT_BLUE, fill=YELLOW, fmt=RM)
    style(ws, "A5", "Prepared by")
    style(ws, "B5", "Your name", font=INPUT_BLUE, fill=YELLOW)

    style(ws, "A7", "CATEGORIES", font=HEAD_FONT, fill=HEAD_FILL)
    style(ws, "B7", "Type", font=HEAD_FONT, fill=HEAD_FILL)
    r = 8
    for cat in INCOME_CATS:
        style(ws, f"A{r}", cat, font=INPUT_BLUE)
        style(ws, f"B{r}", "Income")
        r += 1
    for cat in EXPENSE_CATS:
        style(ws, f"A{r}", cat, font=INPUT_BLUE)
        style(ws, f"B{r}", "Expense")
        r += 1
    cat_first, cat_last = 8, r - 1

    # ------------------------------------------------------ Transactions
    ws = wb.create_sheet("Transactions")
    header_row(ws, 1, ["Date", "Description", "Money in", "Money out", "Category", "Notes"],
               [12, 40, 14, 14, 24, 30])
    # one example row showing the expected format
    style(ws, "A2", "2026-08-01", font=INPUT_BLUE, border=True)
    style(ws, "B2", "EXAMPLE — Cash sales for the day (delete this row)", font=INPUT_BLUE, border=True)
    style(ws, "C2", 850, font=INPUT_BLUE, fmt=RM, border=True)
    style(ws, "D2", None, border=True)
    style(ws, "E2", "Sales", font=INPUT_BLUE, border=True)
    style(ws, "F2", "", border=True)

    dv = DataValidation(
        type="list",
        formula1="=Settings!$A${}:$A${}".format(cat_first, cat_last),
        allow_blank=True, showErrorMessage=True,
        error="Pick a category from the Settings sheet list.")
    ws.add_data_validation(dv)
    dv.add("E2:E%d" % (TXN_ROWS + 1))
    for row in range(3, TXN_ROWS + 2):
        for col in "CD":
            ws[f"{col}{row}"].number_format = RM

    # -------------------------------------------------------------- P&L
    ws = wb.create_sheet("P&L")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 16
    style(ws, "A1", "PROFIT & LOSS", font=HEAD_FONT, fill=HEAD_FILL)
    style(ws, "B1", "='Settings'!B3", font=Font(name=ARIAL, bold=True, color="008000"))
    txn = "Transactions"
    style(ws, "A3", "INCOME", font=BOLD)
    r = 4
    inc_rows = []
    for i, _cat in enumerate(INCOME_CATS):
        style(ws, f"A{r}", "='Settings'!A%d" % (cat_first + i), font=GREEN)
        style(ws, f"B{r}",
              "=SUMIFS('%s'!C2:C%d,'%s'!E2:E%d,A%d)-SUMIFS('%s'!D2:D%d,'%s'!E2:E%d,A%d)"
              % (txn, TXN_ROWS + 1, txn, TXN_ROWS + 1, r,
                 txn, TXN_ROWS + 1, txn, TXN_ROWS + 1, r), fmt=RM)
        inc_rows.append(r)
        r += 1
    style(ws, f"A{r}", "Total income", font=BOLD)
    style(ws, f"B{r}", "=SUM(B%d:B%d)" % (inc_rows[0], inc_rows[-1]), font=BOLD, fmt=RM)
    total_income_row = r

    r += 2
    style(ws, f"A{r}", "EXPENSES", font=BOLD)
    r += 1
    exp_rows = []
    for i, _cat in enumerate(EXPENSE_CATS):
        style(ws, f"A{r}", "='Settings'!A%d" % (cat_first + len(INCOME_CATS) + i), font=GREEN)
        style(ws, f"B{r}",
              "=SUMIFS('%s'!D2:D%d,'%s'!E2:E%d,A%d)-SUMIFS('%s'!C2:C%d,'%s'!E2:E%d,A%d)"
              % (txn, TXN_ROWS + 1, txn, TXN_ROWS + 1, r,
                 txn, TXN_ROWS + 1, txn, TXN_ROWS + 1, r), fmt=RM)
        exp_rows.append(r)
        r += 1
    style(ws, f"A{r}", "Total expenses", font=BOLD)
    style(ws, f"B{r}", "=SUM(B%d:B%d)" % (exp_rows[0], exp_rows[-1]), font=BOLD, fmt=RM)
    total_exp_row = r

    r += 2
    style(ws, f"A{r}", "NET PROFIT / (LOSS)", font=BOLD)
    style(ws, f"B{r}", "=B%d-B%d" % (total_income_row, total_exp_row), font=BOLD, fmt=RM)
    net_row = r

    r += 2
    style(ws, f"A{r}", "Uncategorized transactions", font=PLAIN)
    style(ws, f"B{r}",
          '=COUNTIFS(\'%s\'!A2:A%d,"<>",\'%s\'!E2:E%d,"")' % (txn, TXN_ROWS + 1, txn, TXN_ROWS + 1))
    style(ws, f"A{r + 1}", "(must be 0 before you send the report)", font=Font(name=ARIAL, italic=True, size=9))
    uncat_row = r

    # ------------------------------------------------------ Cash Summary
    ws = wb.create_sheet("Cash Summary")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 16
    style(ws, "A1", "CASH SUMMARY", font=HEAD_FONT, fill=HEAD_FILL)
    style(ws, "A3", "Opening balance")
    style(ws, "B3", "='Settings'!B4", font=GREEN, fmt=RM)
    style(ws, "A4", "Total money in")
    style(ws, "B4", "=SUM('%s'!C2:C%d)" % (txn, TXN_ROWS + 1), fmt=RM)
    style(ws, "A5", "Total money out")
    style(ws, "B5", "=SUM('%s'!D2:D%d)" % (txn, TXN_ROWS + 1), fmt=RM)
    style(ws, "A6", "Closing balance", font=BOLD)
    style(ws, "B6", "=B3+B4-B5", font=BOLD, fmt=RM)

    # ----------------------------------------------------- Client Report
    ws = wb.create_sheet("Client Report")
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 18
    style(ws, "A1", "MONTHLY REPORT", font=Font(name=ARIAL, bold=True, size=16))
    style(ws, "A2", "='Settings'!B2", font=Font(name=ARIAL, size=12, color="008000"))
    style(ws, "A3", "='Settings'!B3", font=Font(name=ARIAL, size=11, color="008000"))

    style(ws, "A5", "THE MONTH IN FOUR NUMBERS", font=HEAD_FONT, fill=HEAD_FILL)
    style(ws, "B5", "", fill=HEAD_FILL)
    style(ws, "A6", "Income")
    style(ws, "B6", "='P&L'!B%d" % total_income_row, font=GREEN, fmt=RM)
    style(ws, "A7", "Expenses")
    style(ws, "B7", "='P&L'!B%d" % total_exp_row, font=GREEN, fmt=RM)
    style(ws, "A8", "Net profit / (loss)", font=BOLD)
    style(ws, "B8", "='P&L'!B%d" % net_row, font=Font(name=ARIAL, bold=True, color="008000"), fmt=RM)
    style(ws, "A9", "Cash in hand at month end")
    style(ws, "B9", "='Cash Summary'!B6", font=GREEN, fmt=RM)

    style(ws, "A11", "TOP 3 EXPENSES", font=HEAD_FILL and HEAD_FONT, fill=HEAD_FILL)
    style(ws, "B11", "", fill=HEAD_FILL)
    exp_range = "'P&L'!B%d:B%d" % (exp_rows[0], exp_rows[-1])
    name_range = "'P&L'!A%d:A%d" % (exp_rows[0], exp_rows[-1])
    for k in range(1, 4):
        row = 11 + k
        style(ws, f"A{row}",
              "=IF(IFERROR(LARGE(%s,%d),0)<=0,\"-\",INDEX(%s,MATCH(LARGE(%s,%d),%s,0)))"
              % (exp_range, k, name_range, exp_range, k, exp_range), font=GREEN)
        style(ws, f"B{row}", "=IFERROR(LARGE(%s,%d),0)" % (exp_range, k), font=GREEN, fmt=RM)

    style(ws, "A16", "NOTES FROM YOUR ACCOUNTANT", font=HEAD_FONT, fill=HEAD_FILL)
    style(ws, "B16", "", fill=HEAD_FILL)
    style(ws, "A17", "Write 2-3 plain-language observations here before sending.",
          font=INPUT_BLUE, fill=YELLOW)
    ws.row_dimensions[17].height = 60
    ws["A17"].alignment = Alignment(wrap_text=True, vertical="top")
    style(ws, "A19", "Prepared by:")
    style(ws, "B19", "='Settings'!B5", font=GREEN)
    style(ws, "A20", "Uncategorized items check")
    style(ws, "B20", "='P&L'!B%d" % uncat_row, font=GREEN)

    wb.save(OUT)
    print("wrote", OUT)


if __name__ == "__main__":
    main()
