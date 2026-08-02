#!/usr/bin/env python3
"""Build all digital products for the store.

Outputs:
  products/budget-tracker/Budget-and-Savings-Tracker.xlsx
  products/invoice-kit/Freelancer-Invoice-Kit.xlsx
  products/focus-planner/Focus-Planner-A4.pdf
  products/focus-planner/Focus-Planner-Letter.pdf
  dist/*.zip  (upload-ready packages, one per product)

Run:  python3 scripts/build_products.py
Then: python3 <xlsx-skill>/scripts/recalc.py on each .xlsx (done by CI/dev, not this script)
"""

import os
import zipfile

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# ---------------------------------------------------------------- styling ---
ARIAL = "Arial"
ACCENT = "2A6F6A"          # muted teal
ACCENT_LIGHT = "E4F0EF"
YELLOW = "FFF2CC"          # fill-me-in cells
GREY = "F2F2F2"
BLUE_FONT = Font(name=ARIAL, color="0000FF")           # hardcoded inputs
BOLD = Font(name=ARIAL, bold=True)
WHITE_BOLD = Font(name=ARIAL, bold=True, color="FFFFFF")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def base_font(ws):
    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(name=ARIAL)


def header_cell(ws, coord, text):
    c = ws[coord]
    c.value = text
    c.font = WHITE_BOLD
    c.fill = PatternFill("solid", fgColor=ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = BOX
    return c


def input_cell(ws, coord, value=None, num_fmt=None):
    c = ws[coord]
    if value is not None:
        c.value = value
    c.font = BLUE_FONT
    c.fill = PatternFill("solid", fgColor=YELLOW)
    c.border = BOX
    if num_fmt:
        c.number_format = num_fmt
    return c


MONEY = "#,##0.00;(#,##0.00);-"
PCT = "0.0%"

# ------------------------------------------------------- product 1: budget --

BUDGET_CATEGORIES = [
    "Housing / Rent", "Utilities", "Groceries", "Transport", "Insurance",
    "Health", "Dining Out", "Entertainment", "Subscriptions", "Shopping",
    "Education", "Other",
]


def build_budget_tracker(path):
    wb = Workbook()

    # --- Start Here ---
    ws = wb.active
    ws.title = "Start Here"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 95
    ws["B2"] = "Budget & Savings Tracker"
    ws["B2"].font = Font(name=ARIAL, bold=True, size=18, color=ACCENT)
    lines = [
        "",
        "How to use this workbook (5 minutes to set up):",
        "1.  Budget sheet — type your planned monthly income and planned spend per category in the yellow cells.",
        "2.  Expense Log sheet — each time you spend money, add one row: date, category, description, amount.",
        "    Pick the category from the list on the Budget sheet so totals match automatically.",
        "3.  Savings Goals sheet — list each goal, its target amount, and what you have saved so far.",
        "",
        "Everything else calculates itself: actual spend per category, money left, net for the month,",
        "your savings rate, and progress toward every goal.",
        "",
        "Legend:",
        "•  Yellow cells with blue text = type here. Everything else is a formula — leave it alone.",
        "•  The workbook is currency-neutral: numbers are plain amounts, so it works with any currency.",
        "•  One realistic example row is pre-filled in each sheet to show the expected format. Overwrite it with",
        "    your own numbers when you start.",
        "",
        "Tip: duplicate the whole file each month (File > Save a Copy) to keep a monthly archive.",
    ]
    for i, line in enumerate(lines, start=3):
        ws[f"B{i}"] = line
        ws[f"B{i}"].font = Font(name=ARIAL, size=11)

    # --- Budget ---
    ws = wb.create_sheet("Budget")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    for col, w in zip("BCDE", (24, 14, 14, 14)):
        ws.column_dimensions[col].width = w

    ws["B2"] = "Monthly Budget"
    ws["B2"].font = Font(name=ARIAL, bold=True, size=14, color=ACCENT)

    # Income block
    ws["B4"] = "INCOME"
    ws["B4"].font = BOLD
    for coord, text in (("B5", "Source"), ("C5", "Planned"),
                        ("D5", "Actual"), ("E5", "Difference")):
        header_cell(ws, coord, text)
    income_rows = ["Salary", "Side income", "Other income"]
    r = 6
    for i, name in enumerate(income_rows):
        row = r + i
        ws[f"B{row}"] = name
        ws[f"B{row}"].border = BOX
        input_cell(ws, f"C{row}", 0, MONEY)
        input_cell(ws, f"D{row}", 0, MONEY)
        ws[f"E{row}"] = f"=D{row}-C{row}"
        ws[f"E{row}"].number_format = MONEY
        ws[f"E{row}"].border = BOX
    # example values in the first income row
    ws["C6"] = 3200
    ws["D6"] = 3200
    inc_first, inc_last = r, r + len(income_rows) - 1
    tot = inc_last + 1
    ws[f"B{tot}"] = "Total income"
    ws[f"B{tot}"].font = BOLD
    for col in "CDE":
        ws[f"{col}{tot}"] = f"=SUM({col}{inc_first}:{col}{inc_last})"
        ws[f"{col}{tot}"].number_format = MONEY
        ws[f"{col}{tot}"].font = BOLD
        ws[f"{col}{tot}"].fill = PatternFill("solid", fgColor=ACCENT_LIGHT)
    income_total_row = tot

    # Expenses block
    start = tot + 2
    ws[f"B{start}"] = "EXPENSES"
    ws[f"B{start}"].font = BOLD
    hdr = start + 1
    for coord, text in ((f"B{hdr}", "Category"), (f"C{hdr}", "Planned"),
                        (f"D{hdr}", "Actual (auto)"), (f"E{hdr}", "Left to spend")):
        header_cell(ws, coord, text)
    exp_first = hdr + 1
    for i, cat in enumerate(BUDGET_CATEGORIES):
        row = exp_first + i
        ws[f"B{row}"] = cat
        ws[f"B{row}"].border = BOX
        input_cell(ws, f"C{row}", 0, MONEY)
        ws[f"D{row}"] = (
            f"=SUMIFS('Expense Log'!$D$2:$D$500,'Expense Log'!$B$2:$B$500,B{row})"
        )
        ws[f"D{row}"].number_format = MONEY
        ws[f"D{row}"].border = BOX
        ws[f"E{row}"] = f"=C{row}-D{row}"
        ws[f"E{row}"].number_format = MONEY
        ws[f"E{row}"].border = BOX
    ws[f"C{exp_first}"] = 1100  # example planned housing
    exp_last = exp_first + len(BUDGET_CATEGORIES) - 1
    tot = exp_last + 1
    ws[f"B{tot}"] = "Total expenses"
    ws[f"B{tot}"].font = BOLD
    for col in "CDE":
        ws[f"{col}{tot}"] = f"=SUM({col}{exp_first}:{col}{exp_last})"
        ws[f"{col}{tot}"].number_format = MONEY
        ws[f"{col}{tot}"].font = BOLD
        ws[f"{col}{tot}"].fill = PatternFill("solid", fgColor=ACCENT_LIGHT)
    exp_total_row = tot

    # Summary block
    s = tot + 2
    ws[f"B{s}"] = "SUMMARY"
    ws[f"B{s}"].font = BOLD
    items = [
        ("Actual income", f"=D{income_total_row}", MONEY),
        ("Actual expenses", f"=D{exp_total_row}", MONEY),
        ("Net this month", f"=D{income_total_row}-D{exp_total_row}", MONEY),
        ("Savings rate",
         f"=IFERROR((D{income_total_row}-D{exp_total_row})/D{income_total_row},0)",
         PCT),
    ]
    for i, (label, formula, fmt) in enumerate(items, start=1):
        row = s + i
        ws[f"B{row}"] = label
        ws[f"B{row}"].border = BOX
        ws[f"C{row}"] = formula
        ws[f"C{row}"].number_format = fmt
        ws[f"C{row}"].font = BOLD
        ws[f"C{row}"].border = BOX
        ws[f"C{row}"].fill = PatternFill("solid", fgColor=ACCENT_LIGHT)

    # --- Expense Log ---
    ws = wb.create_sheet("Expense Log")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCD", (14, 22, 34, 14)):
        ws.column_dimensions[col].width = w
    for coord, text in (("A1", "Date"), ("B1", "Category"),
                        ("C1", "Description"), ("D1", "Amount")):
        header_cell(ws, coord, text)
    # one example row
    input_cell(ws, "A2", "2026-01-03")
    input_cell(ws, "B2", "Groceries")
    input_cell(ws, "C2", "Weekly grocery run")
    input_cell(ws, "D2", 82.40, MONEY)
    ws["F1"] = "Add one row per expense. Category must match a Budget-sheet category exactly."
    ws["F1"].font = Font(name=ARIAL, italic=True, size=9, color="808080")
    for row in range(3, 41):  # pre-styled empty rows for convenience
        for col in "ABCD":
            c = ws[f"{col}{row}"]
            c.border = BOX
            c.font = BLUE_FONT
            if col == "D":
                c.number_format = MONEY

    # --- Savings Goals ---
    ws = wb.create_sheet("Savings Goals")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDE", (26, 14, 14, 14, 12)):
        ws.column_dimensions[col].width = w
    for coord, text in (("A1", "Goal"), ("B1", "Target"), ("C1", "Saved so far"),
                        ("D1", "Remaining"), ("E1", "Progress")):
        header_cell(ws, coord, text)
    input_cell(ws, "A2", "Emergency fund")
    input_cell(ws, "B2", 5000, MONEY)
    input_cell(ws, "C2", 1250, MONEY)
    for row in range(2, 12):
        if row > 2:
            for col in "ABC":
                c = ws[f"{col}{row}"]
                c.border = BOX
                c.font = BLUE_FONT
                if col in "BC":
                    c.number_format = MONEY
        ws[f"D{row}"] = f"=IF(B{row}=\"\",\"\",B{row}-C{row})"
        ws[f"D{row}"].number_format = MONEY
        ws[f"D{row}"].border = BOX
        ws[f"E{row}"] = f"=IF(B{row}=\"\",\"\",IFERROR(C{row}/B{row},0))"
        ws[f"E{row}"].number_format = PCT
        ws[f"E{row}"].border = BOX

    wb.save(path)


# ------------------------------------------------- product 2: invoice kit --

def build_invoice_kit(path):
    wb = Workbook()

    # --- Start Here ---
    ws = wb.active
    ws.title = "Start Here"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 95
    ws["B2"] = "Freelancer Invoice Kit"
    ws["B2"].font = Font(name=ARIAL, bold=True, size=18, color=ACCENT)
    lines = [
        "",
        "What's inside:",
        "•  Invoice — a clean, professional invoice. Fill the yellow cells, print or export to PDF, send.",
        "•  Invoice Log — track every invoice and see paid vs outstanding totals automatically.",
        "•  Clients — a simple client contact list.",
        "",
        "Legend: yellow cells with blue text = type here. Everything else is a formula — leave it alone.",
        "The kit is currency-neutral: amounts are plain numbers, so it works with any currency.",
        "One realistic example row is pre-filled in each sheet to show the expected format — overwrite it",
        "with your own data.",
        "",
        "Tip: duplicate the Invoice sheet for each new invoice (right-click the tab > Move or Copy).",
    ]
    for i, line in enumerate(lines, start=3):
        ws[f"B{i}"] = line
        ws[f"B{i}"].font = Font(name=ARIAL, size=11)

    # --- Invoice ---
    ws = wb.create_sheet("Invoice")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    for col, w in zip("BCDE", (40, 10, 12, 14)):
        ws.column_dimensions[col].width = w

    ws["B2"] = "INVOICE"
    ws["B2"].font = Font(name=ARIAL, bold=True, size=22, color=ACCENT)

    ws["B4"] = "From:"
    ws["B4"].font = BOLD
    input_cell(ws, "B5", "Your Name / Business Name")
    input_cell(ws, "B6", "Your address")
    input_cell(ws, "B7", "your@email.com")

    ws["D4"] = "Invoice #"
    ws["D4"].font = BOLD
    input_cell(ws, "E4", "INV-001")
    ws["D5"] = "Date"
    ws["D5"].font = BOLD
    input_cell(ws, "E5", "2026-01-15")
    ws["D6"] = "Due date"
    ws["D6"].font = BOLD
    input_cell(ws, "E6", "2026-01-29")

    ws["B9"] = "Bill to:"
    ws["B9"].font = BOLD
    input_cell(ws, "B10", "Client name")
    input_cell(ws, "B11", "Client company")
    input_cell(ws, "B12", "client@email.com")

    hdr = 14
    for coord, text in ((f"B{hdr}", "Description"), (f"C{hdr}", "Qty"),
                        (f"D{hdr}", "Rate"), (f"E{hdr}", "Amount")):
        header_cell(ws, coord, text)
    first, last = hdr + 1, hdr + 8
    for row in range(first, last + 1):
        for col, fmt in (("B", None), ("C", "0.##"), ("D", MONEY)):
            c = ws[f"{col}{row}"]
            c.border = BOX
            c.font = BLUE_FONT
            c.fill = PatternFill("solid", fgColor=YELLOW)
            if fmt:
                c.number_format = fmt
        ws[f"E{row}"] = f"=IF(C{row}=\"\",\"\",C{row}*D{row})"
        ws[f"E{row}"].number_format = MONEY
        ws[f"E{row}"].border = BOX
    ws[f"B{first}"] = "Design work — homepage"
    ws[f"C{first}"] = 10
    ws[f"D{first}"] = 60

    r = last + 1
    ws[f"D{r}"] = "Subtotal"
    ws[f"D{r}"].font = BOLD
    ws[f"E{r}"] = f"=SUM(E{first}:E{last})"
    ws[f"E{r}"].number_format = MONEY
    ws[f"E{r}"].font = BOLD
    r += 1
    ws[f"D{r}"] = "Tax rate"
    ws[f"D{r}"].font = BOLD
    input_cell(ws, f"E{r}", 0, PCT)
    tax_rate_row = r
    r += 1
    ws[f"D{r}"] = "Tax"
    ws[f"D{r}"].font = BOLD
    ws[f"E{r}"] = f"=E{last + 1}*E{tax_rate_row}"
    ws[f"E{r}"].number_format = MONEY
    r += 1
    ws[f"D{r}"] = "TOTAL"
    ws[f"D{r}"].font = Font(name=ARIAL, bold=True, size=12)
    ws[f"E{r}"] = f"=E{last + 1}+E{r - 1}"
    ws[f"E{r}"].number_format = MONEY
    ws[f"E{r}"].font = Font(name=ARIAL, bold=True, size=12)
    ws[f"E{r}"].fill = PatternFill("solid", fgColor=ACCENT_LIGHT)

    r += 2
    ws[f"B{r}"] = "Payment details / notes:"
    ws[f"B{r}"].font = BOLD
    input_cell(ws, f"B{r + 1}", "Bank / PayPal / payment link details here")

    # --- Invoice Log ---
    ws = wb.create_sheet("Invoice Log")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEF", (12, 24, 14, 14, 14, 12)):
        ws.column_dimensions[col].width = w
    for coord, text in (("A1", "Invoice #"), ("B1", "Client"), ("C1", "Issued"),
                        ("D1", "Due"), ("E1", "Amount"), ("F1", "Status")):
        header_cell(ws, coord, text)
    input_cell(ws, "A2", "INV-001")
    input_cell(ws, "B2", "Client name")
    input_cell(ws, "C2", "2026-01-15")
    input_cell(ws, "D2", "2026-01-29")
    input_cell(ws, "E2", 600, MONEY)
    input_cell(ws, "F2", "Unpaid")
    for row in range(3, 41):
        for col in "ABCDEF":
            c = ws[f"{col}{row}"]
            c.border = BOX
            c.font = BLUE_FONT
            if col == "E":
                c.number_format = MONEY
    ws["H1"] = "Status must be exactly: Paid or Unpaid"
    ws["H1"].font = Font(name=ARIAL, italic=True, size=9, color="808080")
    summary = [
        ("Total invoiced", "=SUM(E2:E500)"),
        ("Paid", "=SUMIFS(E2:E500,F2:F500,\"Paid\")"),
        ("Outstanding", "=SUMIFS(E2:E500,F2:F500,\"Unpaid\")"),
    ]
    for i, (label, formula) in enumerate(summary):
        row = 3 + i
        ws[f"H{row}"] = label
        ws[f"H{row}"].font = BOLD
        ws[f"I{row}"] = formula
        ws[f"I{row}"].number_format = MONEY
        ws[f"I{row}"].font = BOLD
        ws[f"I{row}"].fill = PatternFill("solid", fgColor=ACCENT_LIGHT)
    ws.column_dimensions["H"].width = 16
    ws.column_dimensions["I"].width = 14

    # --- Clients ---
    ws = wb.create_sheet("Clients")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCD", (24, 20, 28, 40)):
        ws.column_dimensions[col].width = w
    for coord, text in (("A1", "Client"), ("B1", "Contact person"),
                        ("C1", "Email"), ("D1", "Notes")):
        header_cell(ws, coord, text)
    input_cell(ws, "A2", "Acme Studio")
    input_cell(ws, "B2", "Jamie Lee")
    input_cell(ws, "C2", "jamie@acme.example")
    input_cell(ws, "D2", "Net-14 payment terms")
    for row in range(3, 31):
        for col in "ABCD":
            c = ws[f"{col}{row}"]
            c.border = BOX
            c.font = BLUE_FONT

    wb.save(path)


# ------------------------------------------------ product 3: focus planner --

def build_planner(path, pagesize):
    from reportlab.pdfgen import canvas
    from reportlab.lib.colors import HexColor

    W, H = pagesize
    accent = HexColor("#2A6F6A")
    ink = HexColor("#26282B")
    grey = HexColor("#9AA0A6")
    light = HexColor("#D9DDE0")
    M = 50  # margin

    c = canvas.Canvas(path, pagesize=pagesize)
    c.setTitle("The Focus Planner — Undated Habit & Goal Planner")
    c.setAuthor("Clarity Templates")

    def heading(title, subtitle=None):
        c.setFillColor(accent)
        c.rect(M, H - 66, 36, 5, stroke=0, fill=1)
        c.setFillColor(ink)
        c.setFont("Helvetica-Bold", 24)
        c.drawString(M, H - 100, title)
        if subtitle:
            c.setFillColor(grey)
            c.setFont("Helvetica", 11)
            c.drawString(M, H - 118, subtitle)
        c.setFillColor(ink)

    def ruled_lines(x, y_top, width, n, gap=26):
        c.setStrokeColor(light)
        c.setLineWidth(0.8)
        for i in range(n):
            y = y_top - i * gap
            c.line(x, y, x + width, y)
        return y_top - (n - 1) * gap

    def label(x, y, text, size=11, bold=True, color=None):
        c.setFillColor(color or ink)
        c.setFont("Helvetica-Bold" if bold else "Helvetica", size)
        c.drawString(x, y, text)
        c.setFillColor(ink)

    def footer(page_name):
        c.setFillColor(grey)
        c.setFont("Helvetica", 8)
        c.drawString(M, 28, "The Focus Planner")
        c.drawRightString(W - M, 28, page_name)
        c.setFillColor(ink)

    # -- cover --
    c.setFillColor(HexColor("#F4F7F6"))
    c.rect(0, 0, W, H, stroke=0, fill=1)
    c.setFillColor(accent)
    c.rect(0, H - 14, W, 14, stroke=0, fill=1)
    c.rect(0, 0, W, 14, stroke=0, fill=1)
    c.setFillColor(ink)
    c.setFont("Helvetica-Bold", 40)
    c.drawCentredString(W / 2, H * 0.60, "The Focus Planner")
    c.setFillColor(grey)
    c.setFont("Helvetica", 15)
    c.drawCentredString(W / 2, H * 0.60 - 30, "Undated habit & goal planner — start any day, reuse forever")
    c.setStrokeColor(accent)
    c.setLineWidth(2)
    c.line(W / 2 - 60, H * 0.60 - 55, W / 2 + 60, H * 0.60 - 55)
    c.setFillColor(grey)
    c.setFont("Helvetica", 10)
    c.drawCentredString(W / 2, H * 0.25, "Print any pages, as many times as you like, for personal use.")
    c.showPage()

    # -- my goals --
    heading("My Goals", "The three things that matter most this year — and why.")
    y = H - 160
    for i in range(1, 4):
        label(M, y, f"Goal {i}", 13, color=accent)
        ruled_lines(M, y - 26, W - 2 * M, 3)
        label(M, y - 104, "Why it matters:", 10, bold=False, color=grey)
        ruled_lines(M + 90, y - 104, W - 2 * M - 90, 1)
        y -= 170
    footer("Yearly goals")
    c.showPage()

    # -- monthly goals --
    heading("This Month", "Month: ____________________")
    y = H - 160
    label(M, y, "Top 3 outcomes for this month", 13, color=accent)
    ruled_lines(M, y - 26, W - 2 * M, 3, gap=30)
    y -= 130
    label(M, y, "Action steps", 13, color=accent)
    y_top = y - 26
    c.setStrokeColor(light)
    for i in range(8):
        yy = y_top - i * 30
        c.setLineWidth(0.9)
        c.rect(M, yy - 4, 10, 10, stroke=1, fill=0)
        c.line(M + 20, yy, W - M, yy)
    y = y_top - 7 * 30 - 50
    label(M, y, "One habit I'm building this month:", 11, bold=False, color=grey)
    c.setStrokeColor(light)
    c.line(M + 190, y, W - M, y)
    footer("Monthly goals")
    c.showPage()

    # -- habit tracker grid --
    heading("Habit Tracker", "Month: ______________   Mark each day you complete the habit.")
    grid_top = H - 150
    name_w = 130
    days = 31
    rows = 12
    cell_w = (W - 2 * M - name_w) / days
    cell_h = 30
    c.setFont("Helvetica", 6.5)
    c.setFillColor(grey)
    for d in range(days):
        x = M + name_w + d * cell_w + cell_w / 2
        c.drawCentredString(x, grid_top + 5, str(d + 1))
    c.setFillColor(ink)
    c.setStrokeColor(light)
    c.setLineWidth(0.7)
    for r in range(rows + 1):
        y = grid_top - r * cell_h
        c.line(M, y, W - M, y)
    c.line(M, grid_top, M, grid_top - rows * cell_h)
    for d in range(days + 1):
        x = M + name_w + d * cell_w
        c.line(x, grid_top, x, grid_top - rows * cell_h)
    c.setFillColor(grey)
    c.setFont("Helvetica", 9)
    c.drawString(M + 6, grid_top + 5, "Habit")
    footer("Habit tracker")
    c.showPage()

    # -- weekly planner --
    heading("This Week", "Week of: ____________________")
    col_w = (W - 2 * M - 20) / 2
    box_h = (H - 220 - 60) / 4
    day_names = ["Monday", "Tuesday", "Wednesday", "Thursday",
                 "Friday", "Saturday", "Sunday", "Top priorities"]
    for i, name in enumerate(day_names):
        col = i % 2
        row = i // 2
        x = M + col * (col_w + 20)
        y_top = H - 150 - row * (box_h + 12)
        c.setStrokeColor(light)
        c.setLineWidth(1)
        c.roundRect(x, y_top - box_h, col_w, box_h, 6, stroke=1, fill=0)
        is_prio = name == "Top priorities"
        label(x + 12, y_top - 20, name, 11,
              color=accent if is_prio else ink)
        c.setStrokeColor(light)
        c.setLineWidth(0.7)
        for j in range(3):
            yy = y_top - 42 - j * 22
            if yy > y_top - box_h + 10:
                c.line(x + 12, yy, x + col_w - 12, yy)
    footer("Weekly planner")
    c.showPage()

    # -- reflection --
    heading("Monthly Reflection", "Ten minutes at the end of the month keeps the next one on track.")
    y = H - 160
    sections = [
        "Wins — what went well?",
        "Lessons — what didn't, and why?",
        "Next month — what changes?",
    ]
    for title in sections:
        label(M, y, title, 13, color=accent)
        ruled_lines(M, y - 26, W - 2 * M, 4)
        y -= 170
    footer("Reflection")
    c.showPage()

    c.save()


# ---------------------------------------------------------------- packaging --

LICENSE_TEXT = """Thank you for your purchase!

LICENSE — PERSONAL USE
You may use this product for your own personal or business use, on any of
your devices, and print it as many times as you like.

You may NOT resell, redistribute, share, or upload this file (or edited
versions of it) anywhere, free or paid.

Questions or problems with your file? Reply to your purchase receipt email
and we'll make it right.
"""


def make_zip(zip_path, files):
    os.makedirs(os.path.dirname(zip_path), exist_ok=True)
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as z:
        for src, arcname in files:
            z.write(src, arcname)


def main():
    out = {
        "budget": os.path.join(ROOT, "products", "budget-tracker"),
        "invoice": os.path.join(ROOT, "products", "invoice-kit"),
        "planner": os.path.join(ROOT, "products", "focus-planner"),
    }
    for d in out.values():
        os.makedirs(d, exist_ok=True)

    budget_x = os.path.join(out["budget"], "Budget-and-Savings-Tracker.xlsx")
    invoice_x = os.path.join(out["invoice"], "Freelancer-Invoice-Kit.xlsx")
    build_budget_tracker(budget_x)
    build_invoice_kit(invoice_x)

    from reportlab.lib.pagesizes import A4, letter
    planner_a4 = os.path.join(out["planner"], "Focus-Planner-A4.pdf")
    planner_lt = os.path.join(out["planner"], "Focus-Planner-Letter.pdf")
    build_planner(planner_a4, A4)
    build_planner(planner_lt, letter)

    license_path = os.path.join(ROOT, "products", "License.txt")
    with open(license_path, "w") as f:
        f.write(LICENSE_TEXT)

    dist = os.path.join(ROOT, "dist")
    make_zip(os.path.join(dist, "Budget-and-Savings-Tracker.zip"), [
        (budget_x, "Budget-and-Savings-Tracker.xlsx"),
        (license_path, "License.txt"),
    ])
    make_zip(os.path.join(dist, "Freelancer-Invoice-Kit.zip"), [
        (invoice_x, "Freelancer-Invoice-Kit.xlsx"),
        (license_path, "License.txt"),
    ])
    make_zip(os.path.join(dist, "The-Focus-Planner.zip"), [
        (planner_a4, "Focus-Planner-A4.pdf"),
        (planner_lt, "Focus-Planner-US-Letter.pdf"),
        (license_path, "License.txt"),
    ])
    print("Built products and dist zips.")


if __name__ == "__main__":
    main()
